#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Seara Biologia — Fauna
Conversor da planilha de levantamento (.xlsx) para o JSON do app.

Não dá para ler .xlsx em JavaScript puro sem biblioteca, e o app é de zero
dependências por decisão de projeto. Então a leitura da planilha acontece
aqui, uma vez, fora do navegador — e o app importa o JSON.

    python3 importar.py "Cópia de Mineração Planalto.xlsx" -o planalto.json

Depois, no app: Projetos → Importar JSON.

Requer openpyxl (`pip install openpyxl`). É dependência DESTE script, não do
app — o app continua sem nenhuma.

------------------------------------------------------------------------------
MAPEAMENTO (ESPECIFICACAO §5)

  `Inf campo`      linha 2-3 de cada bloco  → unidade amostral + esforço
                   blocos de 46 linhas (10) → registros
  `Autoecologia`   linha 4 = cabeçalho, dados da 5 → táxon + atributos
  `Resumo`         B6:V17                   → metadados das unidades
  `Outros`         J19:K20                  → dados secundários

ATENÇÃO — em `Autoecologia`, ordem e família só aparecem na PRIMEIRA espécie
de cada grupo e ficam `0` nas seguintes. É agrupamento visual do Excel, não
dado ausente: tem que arrastar o último valor válido para baixo.
"""

import argparse
import datetime
import json
import os
import random
import re
import string
import sys
import unicodedata

try:
    import openpyxl
except ImportError:                                        # pragma: no cover
    sys.exit('Falta o openpyxl.  Instale com:  pip3 install openpyxl')


# --------------------------------------------------------------------------- #
# utilidades
# --------------------------------------------------------------------------- #

_ALFABETO = string.ascii_lowercase + string.digits
_CONTADOR = {'n': 0}


def uid(prefixo):
    """Id estável dentro de uma execução, no mesmo formato que o app gera."""
    _CONTADOR['n'] += 1
    sufixo = ''.join(random.choice(_ALFABETO) for _ in range(6))
    return '%s%s%s' % (prefixo, format(_CONTADOR['n'], 'x').rjust(4, '0'), sufixo)


def vazio(v):
    """`0` na planilha do João quer dizer 'em branco' — é preenchimento visual."""
    return v is None or v == 0 or (isinstance(v, str) and not v.strip())


def texto(v):
    if vazio(v):
        return ''
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, datetime.date):
        return v.isoformat()
    if isinstance(v, datetime.time):
        return v.strftime('%H:%M')
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def numero(v):
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(',', '.'))
    except ValueError:
        return None


def data_iso(v):
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, datetime.date):
        return v.isoformat()
    return ''


def hora_hm(v):
    if isinstance(v, datetime.time):
        return '' if (v.hour == 0 and v.minute == 0) else v.strftime('%H:%M')
    if isinstance(v, datetime.datetime):
        return '' if (v.hour == 0 and v.minute == 0) else v.strftime('%H:%M')
    return ''


def normalizar(s):
    s = unicodedata.normalize('NFD', str(s or '').strip().lower())
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return re.sub(r'\s+', ' ', s)


def partir_nome(nome):
    """Quebra 'Tyrannus melancholicus' sem tocar na caixa das letras."""
    partes = re.sub(r'\s+', ' ', str(nome or '').strip()).split(' ')
    return (partes[0] if partes else '', ' '.join(partes[1:]))


# --------------------------------------------------------------------------- #
# leitura da planilha
# --------------------------------------------------------------------------- #

BLOCO = 46          # cada transecto ocupa 46 linhas em `Inf campo`
N_BLOCOS = 10
PRIMEIRA_ESPECIE = 4   # deslocamento da 1ª linha de espécie dentro do bloco
LINHAS_ESPECIE = 40

# Colunas de `Autoecologia`: chave do atributo no app → coluna da planilha
COLUNAS_AUTOECOLOGIA = [
    ('iucn', 6),
    ('listaNacional', 7),      # 'Port 444'  — ver aviso legal no README
    ('listaEstadual', 8),      # 'COPAM 147' — Minas Gerais
    ('sensibilidade', 9),
    ('endemismo', 10),
    ('dependenciaMata', 11),
    ('habitat', 12),
    ('dieta', 13),
]

# A planilha grava alguns valores com caixa inconsistente ('B' e 'b', 'Cam' e
# 'cam'). O domínio do app é minúsculo; a normalização é só de caixa, e só
# nestes campos codificados — jamais no nome científico.
DOMINIOS_MINUSCULOS = {'sensibilidade', 'habitat', 'dieta', 'migratoria',
                       'cinegetica', 'exotica'}

# Siglas de lista vermelha são maiúsculas por convenção (VU, EN, CR, LC). A
# planilha traz 'en' e 'vu' minúsculos em algumas linhas — e era isso que fazia
# a Amazona vinacea, ameaçada, ficar de fora das espécies de interesse.
DOMINIOS_MAIUSCULOS = {'iucn', 'listaNacional', 'listaEstadual'}


def ler_taxons(wb, aba='Autoecologia', avisos=None):
    """Catálogo de espécies com autoecologia. Arrasta ordem e família."""
    if aba not in wb.sheetnames:
        return []
    ws = wb[aba]
    ordem_atual, familia_atual = '', ''
    saida = []
    for r in range(5, ws.max_row + 1):
        bruto = ws.cell(r, 4).value
        # A aba termina com uma linha de totais (14 ordens · 25 famílias ·
        # 59 espécies) que cai exatamente nesta coluna. Nome de espécie tem
        # letra; total é número.
        if not isinstance(bruto, str) or not re.search(r'[A-Za-zÀ-ÿ]', bruto):
            continue
        especie = texto(bruto)
        if not especie:
            continue
        o = texto(ws.cell(r, 2).value)
        f = texto(ws.cell(r, 3).value)
        if o:
            ordem_atual = o           # ← o arrasto: só a 1ª linha do grupo traz
        if f:
            familia_atual = f
        genero, epiteto = partir_nome(especie)

        atributos = {}
        for chave, col in COLUNAS_AUTOECOLOGIA:
            v = ws.cell(r, col).value
            if vazio(v):
                continue
            s = texto(v)
            if chave in DOMINIOS_MINUSCULOS:
                s = s.lower()
            elif chave in DOMINIOS_MAIUSCULOS:
                s = s.upper()
            atributos[chave] = s

        saida.append({
            'id': uid('t'),
            'grupo': 'aves',
            'reino': 'Animalia',
            'filo': 'Chordata',
            'classe': 'Aves',
            'ordem': ordem_atual,
            'familia': familia_atual,
            'genero': genero,
            'epiteto': epiteto,
            'nomeComum': texto(ws.cell(r, 5).value),
            'autor': '',
            'statusTaxonomico': '',
            'chaveGbif': '',
            'sinonimoDe': '',
            'foraDoCatalogo': False,
            'atributos': atributos,
            'gbif': None,
        })
    return saida


def ler_blocos(wb, aba):
    """Os 10 blocos de `Inf campo`: cabeçalho da unidade + registros."""
    if aba not in wb.sheetnames:
        return []
    ws = wb[aba]
    blocos = []
    for k in range(N_BLOCOS):
        base = 2 + BLOCO * k                     # linha do cabeçalho do bloco
        cab = base + 1                           # linha de dados do cabeçalho
        info = {
            'numero': texto(ws.cell(cab, 2).value) or str(k + 1),
            'data': data_iso(ws.cell(cab, 3).value),
            'horaInicio': hora_hm(ws.cell(cab, 4).value),
            'horaFim': hora_hm(ws.cell(cab, 5).value),
            'nebulosidade': texto(ws.cell(cab, 6).value),
            'latitude': texto(ws.cell(cab, 9).value),
            'longitude': texto(ws.cell(cab, 10).value),
            'comprimentoKm': numero(ws.cell(cab, 11).value),
            'areaHa': numero(ws.cell(cab, 12).value),
        }
        registros = []
        for r in range(base + PRIMEIRA_ESPECIE, base + PRIMEIRA_ESPECIE + LINHAS_ESPECIE):
            especie = texto(ws.cell(r, 5).value)
            if not especie:
                continue
            registros.append({
                'ordem': texto(ws.cell(r, 3).value),
                'familia': texto(ws.cell(r, 4).value),
                'especie': especie,
                'nomeComum': texto(ws.cell(r, 6).value),
                'quantidade': numero(ws.cell(r, 7).value) or 1,
                'foto': texto(ws.cell(r, 8).value),
            })
        blocos.append({'info': info, 'registros': registros})
    return blocos


def ler_resumo(wb):
    """Metadados das unidades em `Resumo` B6:V17 (cada TST ocupa 2 colunas)."""
    if 'Resumo' not in wb.sheetnames:
        return {}, []
    ws = wb['Resumo']
    geral = {
        'cliente': texto(ws.cell(3, 3).value),
        'municipio': texto(ws.cell(3, 10).value),
        'sazonalidade1': texto(ws.cell(3, 16).value).lower(),
        'sazonalidade2': texto(ws.cell(4, 16).value).lower(),
    }
    unidades = []
    for k in range(N_BLOCOS):
        col = 3 + 2 * k                       # coluna da 1ª campanha
        unidades.append({
            'codigo': texto(ws.cell(6, col).value) or ('TST %02d' % (k + 1)),
            'areaInfluencia': texto(ws.cell(7, col).value),
            'comprimentoKm': numero(ws.cell(8, col).value),
            'areaHa': numero(ws.cell(9, col).value),
            'coordenadas': texto(ws.cell(11, col).value),
            'sazonalidade': texto(ws.cell(12, col).value).lower(),
            'data': data_iso(ws.cell(13, col).value),
        })
    return geral, unidades


def ler_secundarios(wb):
    if 'Outros' not in wb.sheetnames:
        return {'titulo': '', 'autores': '', 'especies': []}
    ws = wb['Outros']
    return {
        'titulo': texto(ws.cell(19, 11).value),
        'autores': texto(ws.cell(20, 11).value),
        'especies': [],
    }


# --------------------------------------------------------------------------- #
# montagem do estado do app
# --------------------------------------------------------------------------- #

ABAS_CAMPANHA = [
    ('Inf campo', '1ª campanha'),
    ('Inf campo 2', '2ª campanha'),
]


def converter(caminho, nome_projeto=None, largura_m=50.0):
    wb = openpyxl.load_workbook(caminho, data_only=True)
    avisos = []

    taxons = ler_taxons(wb, avisos=avisos)
    por_nome = {normalizar('%s %s' % (t['genero'], t['epiteto'])): t for t in taxons}
    geral, meta_unidades = ler_resumo(wb)

    projeto = {
        'id': uid('p'),
        'nome': nome_projeto or os.path.splitext(os.path.basename(caminho))[0],
        'cliente': geral.get('cliente', ''),
        'municipio': geral.get('municipio', ''),
        'uf': 'MG',
        'orgao': '',
        'processo': '',
        'responsavel': '',
        'registroProfissional': '',
        'grupo': 'aves',
        'observacao': 'Importado da planilha %s.' % os.path.basename(caminho),
        'secundarios': ler_secundarios(wb),
        'campanhas': [],
    }

    total_unidades = 0
    total_registros = 0

    for indice, (aba, rotulo) in enumerate(ABAS_CAMPANHA):
        if aba not in wb.sheetnames:
            continue
        saz = geral.get('sazonalidade%d' % (indice + 1), '')
        campanha = {
            'id': uid('c'),
            'rotulo': rotulo,
            'sazonalidade': saz if saz in ('chuva', 'seca') else '',
            'dataInicio': '',
            'dataFim': '',
            'equipe': '',
            'observacao': '',
            'unidades': [],
        }

        blocos = ler_blocos(wb, aba)
        datas = []
        for k, bloco in enumerate(blocos):
            info = bloco['info']
            meta = meta_unidades[k] if k < len(meta_unidades) else {}

            data = info['data'] or (meta.get('data') if indice == 0 else '')
            tem_dado = bool(bloco['registros']) or bool(info['data'])
            if not tem_dado:
                # Bloco só com a moldura: a 2ª campanha da planilha é a 1ª
                # duplicada e vazia. Criar unidade aqui seria inventar campo.
                continue

            comp = info['comprimentoKm'] or meta.get('comprimentoKm')
            esforco = {'larguraM': largura_m}
            if comp:
                esforco['comprimentoKm'] = comp

            unidade = {
                'id': uid('u'),
                'metodo': 'transecto',
                'codigo': meta.get('codigo') or ('TST %02d' % (k + 1)),
                'latitude': info['latitude'],
                'longitude': info['longitude'],
                'altitude': '',
                'fitofisionomia': '',
                'areaInfluencia': meta.get('areaInfluencia') or '',
                'data': data,
                'horaInicio': info['horaInicio'],
                'horaFim': info['horaFim'],
                'nebulosidade': info['nebulosidade'],
                'vento': '',
                'temperatura': '',
                'observacao': '',
                'esforco': esforco,
                'registros': [],
            }
            if data:
                datas.append(data)

            for reg in bloco['registros']:
                chave = normalizar(reg['especie'])
                taxon = por_nome.get(chave)
                if taxon is None:
                    # Nome que aparece na ficha de campo mas não no catálogo de
                    # autoecologia. Não pode virar registro órfão nem ser
                    # descartado em silêncio: entra marcado.
                    genero, epiteto = partir_nome(reg['especie'])
                    taxon = {
                        'id': uid('t'), 'grupo': 'aves', 'reino': 'Animalia',
                        'filo': 'Chordata', 'classe': 'Aves',
                        'ordem': reg['ordem'], 'familia': reg['familia'],
                        'genero': genero, 'epiteto': epiteto,
                        'nomeComum': reg['nomeComum'], 'autor': '',
                        'statusTaxonomico': '', 'chaveGbif': '', 'sinonimoDe': '',
                        'foraDoCatalogo': True, 'atributos': {}, 'gbif': None,
                    }
                    taxons.append(taxon)
                    por_nome[chave] = taxon
                    avisos.append(
                        '“%s” (%s, %s) aparece em %s mas não está na aba '
                        'Autoecologia. Entrou no catálogo marcado como fora do '
                        'catálogo — confira se não é o mesmo bicho com outro nome.'
                        % (reg['especie'], reg['familia'], aba, unidade['codigo']))

                unidade['registros'].append({
                    'id': uid('r'),
                    'taxonId': taxon['id'],
                    'quantidade': reg['quantidade'],
                    'tipo': 'visual',
                    'data': data,
                    'hora': '',
                    'observador': '',
                    'foto': reg['foto'],
                    'observacao': '',
                })
                total_registros += 1

            campanha['unidades'].append(unidade)
            total_unidades += 1

        if datas:
            campanha['dataInicio'] = min(datas)
            campanha['dataFim'] = max(datas)
        projeto['campanhas'].append(campanha)

    # ---- conferências que valem um aviso na tela ---------------------------
    sem_esforco = [u['codigo']
                   for c in projeto['campanhas'] for u in c['unidades']
                   if not u['esforco'].get('comprimentoKm')]
    if sem_esforco:
        avisos.append(
            'A planilha não traz o comprimento de %d transecto(s) (%s). Sem '
            'esforço não há densidade nem comparação entre métodos — preencha '
            'na tela de Unidades amostrais.'
            % (len(sem_esforco), ', '.join(sem_esforco)))

    sem_registro = [t for t in taxons
                    if not any(r['taxonId'] == t['id']
                               for c in projeto['campanhas']
                               for u in c['unidades']
                               for r in u['registros'])]
    if sem_registro:
        avisos.append(
            '%d espécie(s) do catálogo de autoecologia não têm registro em '
            'nenhuma unidade amostral. Na planilha elas vieram de observação '
            'ocasional, fora dos transectos: crie uma unidade de busca ativa '
            'para elas, senão não entram na riqueza nem na curva do coletor.'
            % len(sem_registro))

    vazias = [c['rotulo'] for c in projeto['campanhas'] if not c['unidades']]
    if vazias:
        avisos.append(
            'Campanha(s) sem nenhuma unidade amostral: %s. Na planilha as abas '
            'existem duplicadas mas estão em branco.' % ', '.join(vazias))

    estado = {
        'versao': 1,
        'projetos': [projeto],
        'taxons': taxons,
        'tema': 'auto',
        'projetoAtivo': projeto['id'],
        'campanhaAtiva': projeto['campanhas'][0]['id'] if projeto['campanhas'] else None,
        'unidadeAtiva': (projeto['campanhas'][0]['unidades'][0]['id']
                         if projeto['campanhas'] and projeto['campanhas'][0]['unidades'] else None),
        'exemploSemeado': True,
        'importacao': {
            'origem': os.path.basename(caminho),
            'quando': datetime.datetime.now().isoformat(timespec='seconds'),
            'avisos': avisos,
        },
    }

    resumo = {
        'especies': len(taxons),
        'especiesCatalogo': len([t for t in taxons if not t['foraDoCatalogo']]),
        'familias': len({t['familia'] for t in taxons if t['familia']}),
        'ordens': len({t['ordem'] for t in taxons if t['ordem']}),
        'unidades': total_unidades,
        'campanhas': len(projeto['campanhas']),
        'registros': total_registros,
        'individuos': sum(r['quantidade']
                          for c in projeto['campanhas']
                          for u in c['unidades'] for r in u['registros']),
        'avisos': avisos,
    }
    return estado, resumo


def main():
    ap = argparse.ArgumentParser(
        description='Converte a planilha de levantamento de fauna para o JSON do app.')
    ap.add_argument('planilha', help='arquivo .xlsx')
    ap.add_argument('-o', '--saida', default=None, help='arquivo .json de saída')
    ap.add_argument('-n', '--nome', default=None, help='nome do projeto')
    ap.add_argument('--largura', type=float, default=50.0,
                    help='largura da faixa do transecto, em metros (padrão 50)')
    args = ap.parse_args()

    estado, resumo = converter(args.planilha, args.nome, args.largura)

    saida = args.saida or (os.path.splitext(os.path.basename(args.planilha))[0] + '.json')
    with open(saida, 'w', encoding='utf-8') as fh:
        json.dump(estado, fh, ensure_ascii=False, indent=1)

    print('Gravado em %s' % saida)
    print('')
    print('  campanhas .............. %d' % resumo['campanhas'])
    print('  unidades amostrais ..... %d' % resumo['unidades'])
    print('  registros .............. %d' % resumo['registros'])
    print('  indivíduos ............. %d' % resumo['individuos'])
    print('  espécies (catálogo) .... %d' % resumo['especiesCatalogo'])
    print('  espécies (total) ....... %d' % resumo['especies'])
    print('  famílias ............... %d' % resumo['familias'])
    print('  ordens ................. %d' % resumo['ordens'])
    if resumo['avisos']:
        print('')
        print('Avisos:')
        for a in resumo['avisos']:
            print('  · %s' % a)


if __name__ == '__main__':
    main()
